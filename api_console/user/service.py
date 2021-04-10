from django.conf import settings
from openpyxl import Workbook
from datetime import datetime
from openpyxl.utils import get_column_letter
from virtual_day.users.models import User
import os
from virtual_day.utils import constants
from openpyxl.styles import Font
import glob


def export_excel_file() -> str:
    users = User.objects.only(
        'id', 'first_name', 'last_name', 'email', 'address', 'role',
        'phone', 'is_active', 'language'
    ).all()
    workbook = Workbook()

    # grab the active worksheet
    worksheet = workbook.active
    worksheet.title = 'Пользователи'
    # Define the titles for columns
    columns = [
        'ID',
        'Имя',
        'Фамилия',
        'Почта',
        'Адрес',
        'Роль студента',
        'Телефон',
        'Активен',
        'Язык пользователя',
    ]
    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        width = len(column_title) + 8
        if len(column_title) > 8:
            width = len(column_title) + 3
        elif len(column_title) == 5:
            width = len(column_title) + 25
        column_dimensions.width = width
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.font = Font(size=12)
        cell.value = column_title

    for user in users:
        row_num += 1
        user_active_status = "НЕТ"
        if user.is_active:
            user_active_status = "ДА"
        row = [
            user.id,
            user.first_name,
            user.last_name,
            user.email,
            user.address,
            constants.USER_TYPES[user.role][1],
            user.phone,
            user_active_status,
            user.language,
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.font = Font(size=12)
            cell.value = cell_value

    # Save the file
    if not os.path.exists(settings.MEDIA_ROOT + f"/excel_dumps/"):
        os.makedirs(settings.MEDIA_ROOT + f"/excel_dumps/")
    for filename in glob.glob(
            settings.MEDIA_ROOT + f"/excel_dumps/users_*")[:-4]:
        os.remove(filename)
    filename = f"users_{datetime.now().strftime('%d-%m-%Y_%H-%M')}"
    workbook.save(settings.MEDIA_ROOT + f"/excel_dumps/{filename}.xlsx")
    url = settings.SITE_URL + \
          settings.MEDIA_URL + f"excel_dumps/{filename}.xlsx"
    return url
